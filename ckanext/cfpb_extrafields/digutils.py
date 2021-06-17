from collections import namedtuple
import re

from openpyxl import load_workbook

from ckanext.cfpb_extrafields import validators as v

Version = namedtuple("Version", "document_version schema_version date")
def get_versions(workbook):
    versions = []
    try:
        sheet = workbook["Version History"]
    except KeyError:
        return versions
    col = 3
    row = 8
    while sheet.cell(row=row, column=col).value:
        versions.append(Version(*(sheet.cell(row=row, column=col+x).value for x in range(3))))
        row += 1
    return versions

def get_schema_version(workbook):
    versions = get_versions(workbook)
    if not versions:
        return 0
    else:
        return int(max(versions).schema_version)

class NamedWorkbook(object):
    def __init__(self, wb):
        self.wb = wb
    def __getitem__(self, key):
        return get_named_cell(key, self.wb)


def get_named_cell(name, wb):
    sheet, cellname = next(wb.get_named_range(name).destinations)
    cell = wb[sheet][cellname]
    value = cell.value
    if isinstance(value, basestring) and value.startswith("="):
        cell = get_named_cell(value[1:], wb)
    return cell

# Helper functions to convert values of sheets
def strfy(val):
    if isinstance(val, basestring):
        if val.lower().strip() in ["na", "n/a", "not applicable", "select one"]:
            return ""
        return val.strip()
    else:
        return ""

def access_restrictions(supervisor_cell, owner_cell, addl_cell):
    def get_access_restrictions(ws):
        restrictions = []
        for field, cell in [("Supervisor", supervisor_cell), ("Data Owner", owner_cell)]:
            if ws[cell].value == "Yes":
                restrictions.append("{0} approval required".format(field))
        addl_restrictions = strfy(ws[addl_cell].value)
        if addl_restrictions:
            restrictions.append(addl_restrictions)
        return ", ".join(restrictions)
    return get_access_restrictions

# FIELDS values can be functions that take in a worksheet and provide a value
# These factories return functions to use there.
# For instance, pra_exclusion is in fields D38 and D39.
# Therefore, we want the FIELDS value to be a function that takes in a worksheet and returns
# those cells added together.
# Instead of making this function manually with `dig_func = lambda ws: ws["D38"]+ws["D39"]`,
# we can say `dig_func = concat(["D38", "D39"])` and it will make the function and return it to us.
def concat(fields):
    def concat_fields(ws):
        result = ""
        for field in fields:
            result += strfy(ws[field].value)
        return result
    return concat_fields

def date(cell):
    def get_date(ws):
        val = ws[cell].value
        if hasattr(val, "strftime"):
            val = val.strftime("%Y-%m-%d")
        val = strfy(val) # Treat "n/a" as empty string
        _ = v.reasonable_date_validator(val) # Make sure it's a valid date, but return the string.
        return val
    return get_date

def lower(cell):
    def get_lower(ws):
        return (ws[cell].value or "").lower()
    return get_lower

def sub(cell, pattern, replacement):
    def get_sub(ws):
        return re.sub(pattern, replacement, ws[cell].value or "")
    return get_sub

# Strip all non-digit characters from this field
transfer_initial_size = sub("B47", "[^0-9.]", "")

def _any(*cells):
    def get_any(ws):
        return "yes" if any((lower(cell)(ws) == "yes" for cell in cells)) else "no"
    return get_any

# Maps field name to either a cell or a function that's passed the worksheet and should return the value
# Note that some values are currently blank and commented out as they don't map to any fields in the DIG excel sheet
FIELDS_BY_VERSION = {
    0: {
        "access_restrictions": access_restrictions("B16", "D16", "B17"),
        "contact_primary_name": "F16",
        # "contact_secondary_name": "B6",
        "data_source_names": "D10",
        # "dataset_notes": "",
        "dig_id": lambda ws: v.dig_id_validator(strfy(ws["B5"].value)),
        "initial_purpose_for_intake": "H15",
        "legal_authority_for_collection": "B25",
        "notes": "H4",
        "pra_exclusion": "D38",
        "pra_omb_control_number": lambda ws: v.pra_control_num_validator(strfy(ws["F37"].value)),
        "pra_omb_expiration_date": date("F38"),
        "privacy_contains_pii": _any("B29", "F29"),
        "privacy_has_direct_identifiers": _any("B30", "F30"),
        "privacy_has_privacy_act_statement": lower("D30"),
        "privacy_pia_notes": "B33",
        "privacy_pia_title": "D32",
        "privacy_sorn_number": "D31",
        "private": lambda ws: True,# Always have new data sources default to private
        "procurement_document_id": "F24",
        "relevant_governing_documents": "D24",
        "sensitivity_level": "B13",
        "title": "B4",
        "transfer_details": "B54",
        "transfer_initial_size": transfer_initial_size,
        "transfer_method": "B48",
        "update_frequency": "F47",
        "usage_restrictions":  concat(["B18", "B19"]),
        # "website_url": "",
        # "wiki_link": "",
    },
    1: {
        "access_restrictions": access_restrictions("dig_sensitivity_supervisor_approval", "dig_sensitivity_owner_approval", "dig_access_restrictions"),
        "contact_primary_name": "dig_cfpb_poc",
        "contact_primary_email": "dig_email",
        # "contact_secondary_name": "B6",
        "data_source_names": "dig_ext_source", #?
        # "dataset_notes": "",
        "dig_id": lambda ws: v.dig_id_validator(strfy(ws["dig_id"].value)),
        "initial_purpose_for_intake": "dig_intake_purpose",
        "legal_authority_for_collection": "dig_legal_auth",
        "notes": "dig_description",
        "pra_exclusion": "dig_pra_exemption",
        "pra_omb_control_number": lambda ws: v.pra_control_num_validator(strfy(ws["dig_pra_omb_control"].value)),
        "pra_omb_expiration_date": date("dig_pra_expiration"),
        "privacy_contains_pii": _any("dig_privacy_pii_cfpb", "dig_privacy_pii_3rd"), 
        "privacy_has_direct_identifiers": _any("dig_privacy_dpi_cfpb", "dig_privacy_dpi_3rd"),
        "privacy_has_privacy_act_statement": lower("dig_privacy_statement"),
        "privacy_pia_notes": "dig_privacy_notes",
        "privacy_pia_title": "dig_privacy_pia",
        "privacy_sorn_number": "dig_privacy_sorn",
        "private": lambda ws: True,# Always have new data sources default to private
        "procurement_document_id": "dig_legal_docid",
        "relevant_governing_documents": "dig_legal_agreement_type",
        "sensitivity_level": "dig_sensitivity_cfpb",
        "title": "dig_request_title",
        "transfer_details": "dig_storage_notes",
        "transfer_initial_size": sub("dig_storage_size", "[^0-9.]", ""),
        "transfer_method": "dig_transfer_method",
        "update_frequency": "dig_intake_freq",
        "usage_restrictions":  concat(["dig_storage_permission_notes", "dig_storage_notes"]),
        # "website_url": "",
        # "wiki_link": "",
    },
}

def get_field(worksheet, field, fields):
    cell_or_func = fields[field]
    if callable(cell_or_func):
        return cell_or_func(worksheet)
    else:
        return strfy(worksheet[cell_or_func].value)

def make_rec_from_sheet(ws, fields):
    result = {}
    errors = []
    for field in fields:
        try:
            result[field] = get_field(ws, field, fields)
        except NotFound as  err:
            errors.append(field + ": Unable to extract field from workbook - Check for duplicate or undefined cell ranges")
        except (Exception, StopIteration) as  err:
            # Invalid or Not Found respectively
            errors.append(field + ": " + getattr(err, "error", getattr(err, "message", "UKNOWN_ERROR")))
    return result, errors

def make_rec(excel_file):
    wb = load_workbook(excel_file, read_only=True)
    version = get_schema_version(wb)
    fields = FIELDS_BY_VERSION[version]
    if version == 0:
        ws = wb.worksheets[0]
    else:
        ws = NamedWorkbook(wb)
    return make_rec_from_sheet(ws, fields)
